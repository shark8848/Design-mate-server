from MaterialWarehouse import GlassMaterialWarehouse, WallInsulationMaterialWarehouse, WindowFrameMaterialWarehouse
from BuildingElement import *
from apocolib.GPUConfigurator import GPUConfigurator
from BuildingSpaceBase import *
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from apocolib.MQueueManager import MQueueManager
from apocolib.RabbitMQProducer import RabbitMQProducer
from apocolib.MlLogger import mlLogger as ml_logger
import dataSetBaseParamters as ds_bp
from HouseDataGenerator import HouseDataGenerator, CustomJSONEncoder
import pandas as pd
import json
from datetime import datetime
import pdb
import copy
import random
import os
from openpyxl import Workbook, load_workbook
from apocolib.CouchDBPool import couchdb_pool
from apocolib.timing import timing_decorator
import numpy as np
import sys
sys.path.append("..")


# configurator = GPUConfigurator(use_gpu=False, gpu_memory_limit=None)
# configurator = GPUConfigurator(use_gpu=True, gpu_memory_limit=2048)
# configurator.configure_gpu()
# tf.device(configurator.select_device())
# from websocket_server.webSocketServer import external_push_message as expm

# rmqp = RabbitMQProducer(queue_manager=MQueueManager())
gmWarehouse = GlassMaterialWarehouse()
wimWarehouse = WallInsulationMaterialWarehouse()
wfmWarehouse = WindowFrameMaterialWarehouse()
_ss = SpaceStandard()  # 创建空间标准对象


def check_data_type(tensor):
    if tf.is_tensor(tensor):
        data_type = tensor.dtype
        if data_type.is_floating:
            # print("张量中的数据类型为浮点型")
            return "float"
        elif data_type.is_integer:
            # print("张量中的数据类型为整型")
            return "int"
        else:
            # print("张量中的数据类型为对象")
            return "object"
    else:
        # print("传递的数据不是张量")
        return "error tensor"

# 生成数据集


def generate_dataset(size=1000, model='create', dataset_file='./datasets/training_dataset.xlsx'):  # create;reload

    houses_x = []
    houses_y = []

    ds_size = size  # size if size is not None else ds_bp.get_dataset_size()
    print("Preparing to generate the training dataset. It may take a few minutes. Please be patient.")
    if model == 'create':
        houses_creator = HousesCreator(ds_size)
        # 生成一个房屋对象
        houses = houses_creator.make_houses()

        data_dir = f"./houses_json/training_dataset_json/{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}"

        for i, house in enumerate(houses):

            file_serias = datetime.now().strftime("%Y%m%d_%H%M%S%f")[:21]
            house.save_to_json(f"{data_dir}/{file_serias}.json")
            house.save_to_json_cn(f"{data_dir}/{file_serias}_cn.json")
            houses_features, target = house.to_tensor()
            houses_features = tf.cast(houses_features, dtype=tf.float64)
            target = tf.cast(target, dtype=tf.float64)

            # print(" houses_features is tensor ",tf.is_tensor(houses_features), "houses_features shape ",houses_features.shape, "data_type ",check_data_type(houses_features))
            # print(" target is tensor ",tf.is_tensor(target), "target shape ",target.shape, "data_type ",check_data_type(target))

            houses_x.append(houses_features)
            houses_y.append(target)
            sys.stdout.write("\rGenerate dataset: %d/%d houses_features shape: %s data_type is %s target shape %s  data_type is %s " %
                             (i+1, ds_size, houses_features.shape, check_data_type(houses_features), target.shape, check_data_type(target)))
            sys.stdout.flush()
            # time.sleep(0.1)
            # expm("AC2NNetTrainer",f"Generate dataset:{i+1}/{ds_size}")
            # rmqp.publish(f"Generate dataset:{i+1}/{ds_size}")
            # rmqp.close()
        export_to_excel(houses_x, houses_y, size, dataset_file)
        return houses_x, houses_y

    elif model == 'reload':
        houses_x, houses_y = import_from_excel(
            "./datasets/training_dataset.xlsx")
        return houses_x, houses_y
    else:
        print('Invalid create dataset model {model}')
        return [], []


def update_walls(walls, wimWarehouse, wfmWarehouse, gmWarehouse):
    # updated_walls = []

    for wall in walls:
        wall._material = WallInsulation(area=wall._area, warehouse=wimWarehouse,
                                        material_type=wimWarehouse.get_random_material())
        window = copy.deepcopy(wall.get_window())

        if window is not None:
            window_area = min(random.uniform(0.85, 1.15)
                              * window._area, wall._area)
            aspect_ratio = np.random.uniform(0.618, 1.618)
            window_width = np.sqrt(window_area * aspect_ratio)
            window_height = window_area / window_width

            window_width = min(window_width, wall.get_width())
            window_height = min(window_height, wall.get_height())

            wfa_ratio, wf_material = wfmWarehouse.get_best_wf(
                window_width, window_height)

            window = Window(area=window_area, width=window_width,
                            height=window_height, orientation=wall.get_orientation())
            window_frame = WindowFrame(
                area=window_area*wfa_ratio, warehouse=wfmWarehouse, material_type=wf_material)

            window.set_window_frame(window_frame)
            glass = Glass(area=window_area*(1-wfa_ratio), warehouse=gmWarehouse,
                          material_type=gmWarehouse.get_random_material())
            window.set_glass(glass)
            wall.reset_window(window)

        # updated_walls.append(wall)

    return walls


def generate_test_population(num_individuals=50, loadFromFile=None):

    json_data = None
    with open(loadFromFile, 'r') as f:
        json_data = json.load(f)

    houses_x = []
    houses_y = []

    ds_size = num_individuals  # size if size is not None else ds_bp.get_dataset_size()

    i_house = House().json_to_house(json_data)

    for i in range(num_individuals):
        house = copy.deepcopy(i_house)

        # 随机对house 对象中的窗户面积进行修改
        # pdb.set_trace()
        for room in house._rooms:
            for wall in room._walls:

                wall._material = WallInsulation(area=wall._area, warehouse=wimWarehouse,
                                                material_type=wimWarehouse.get_random_material())
                window = copy.deepcopy(wall.get_window())

                if window is not None:

                    window_area = min(random.uniform(
                        0.85, 1.15) * window._area, wall._area)
                    aspect_ratio = np.random.uniform(0.618, 1.618)
                    window_width = np.sqrt(window_area * aspect_ratio)
                    window_height = window_area / window_width

                    if window_width > wall.get_width():
                        window_width = wall.get_width()
                        window_height = window_area / window_width
                    if window_height > wall.get_height():
                        window_height = wall.get_height()
                        window_width = window_area / window_height

                    wfa_ratio, wf_material = wfmWarehouse.get_best_wf(
                        window_width, window_height)
                    # 生成一个窗户对象
                    window = Window(area=window_area, width=window_width,
                                    height=window_height, orientation=wall.get_orientation())
                    window_frame = WindowFrame(
                        area=window_area*wfa_ratio, warehouse=wfmWarehouse, material_type=wf_material)
                    # 生成一个窗框对象，将窗框添加窗户到上
                    window.set_window_frame(window_frame)
                    # 生成一个玻璃对象，将玻璃添加到窗户上
                    glass = Glass(area=window_area*(1-wfa_ratio), warehouse=gmWarehouse,
                                  material_type=gmWarehouse.get_random_material())
                    window.set_glass(glass)
                    wall.reset_window(window)

        house_features, target = house.to_tensor()
        house_features = tf.cast(house_features, dtype=tf.float64)
        target = tf.cast(target, dtype=tf.float64)

        house_features = house_features.numpy()
        target = target.numpy()

        sys.stdout.write("\rGenerate test population: %d/%d houses_features shape: %s target shape %s " %
                         (i+1, num_individuals, house_features.shape, target.shape))
        sys.stdout.flush()

        houses_x.append(house_features)
        houses_y.append(target)

    population_to_exl(houses_x, houses_y)

    return houses_x, houses_y

# json_doc_id couchdb 中的文档id
@timing_decorator
def generate_test_population_v2(rank='[0/0]', num_individuals=50, json_data=None):

    # begin time
    start_time = time.time()

    building_x = []
    building_y = []

    i_biulding = Building().json_to_building(json_data)

    ml_logger.info("Convert json to building Object done")

    for i in range(num_individuals):
        building = copy.deepcopy(i_biulding)
        # 随机选择一层作为测试
        floor = building.get_floor(
            random.randint(0, building.get_floor_num()-1))

        #ml_logger.info("update houses Object in floor/building ...")
        for house in floor.get_houses():
            for room in house._rooms:
                room.set_walls(update_walls(
                    room._walls, wimWarehouse, wfmWarehouse, gmWarehouse))
        #ml_logger.info("update houses Object in floor/building done")

        #ml_logger.info("upate staircases Object in floor/building ...")
        for staircase in floor._staircases:
            staircase.set_walls(update_walls(
                staircase._walls, wimWarehouse, wfmWarehouse, gmWarehouse))
        #ml_logger.info("upate staircases Object in floor/building done")

        #ml_logger.info("update corridor Object in floor/building ...")
        for corridor in floor._corridors:
            corridor.set_walls(update_walls(
                corridor._walls, wimWarehouse, wfmWarehouse, gmWarehouse))
        #ml_logger.info("update corridor Object in floor/building done")

        #ml_logger.info("convert floor to tensor ...")
        building_features, target = floor.to_tensor()  # 将建筑对象转换为张量
        #ml_logger.info("convert floor to tensor done")

        building_features = tf.cast(
            building_features, dtype=tf.float64)  # 将张量转换为浮点数
        target = tf.cast(target, dtype=tf.float64)  # 将张量转换为浮点数

        building_features = building_features.numpy()  # 将张量转换为numpy数组
        target = target.numpy()  # 将张量转换为numpy数组

        #ml_logger.info(
        #    "convert json object to building object and convert to tensor done")

        building_x.append(building_features)
        building_y.append(target)

        end_time = time.time()

        time_cost = end_time - start_time

        sys.stdout.write("\r\n Rank %s Generating test populations v2:[ %d/%d ] building_features shape: %s target shape %s ,time cost %.2f " %
                         (rank,i+1, num_individuals, building_features.shape, target.shape, time_cost))
        sys.stdout.flush()

    sys.stdout.writelines("\n")

    test_population_file = population_to_excel_T(building_x, building_y)
    ml_logger.info(" Rank %s Generate test populations v2 done, save to %s" %
                   (rank,test_population_file))
    
    # 把 building_x 和 building_y 转换为 numpy 数组
    #building_x = np.array(building_x)
    #building_y = np.array(building_y)

    return building_x, building_y

def generate_building(json_data):

    building = Building().json_to_building(json_data)

    floor = building.get_floor(
    random.randint(0, building.get_floor_num()-1))

    #ml_logger.info("update houses Object in floor/building ...")
    for house in floor.get_houses():
        for room in house._rooms:
            room.set_walls(update_walls(
                room._walls, wimWarehouse, wfmWarehouse, gmWarehouse))
    #ml_logger.info("update houses Object in floor/building done")

    #ml_logger.info("upate staircases Object in floor/building ...")
    for staircase in floor._staircases:
        staircase.set_walls(update_walls(
            staircase._walls, wimWarehouse, wfmWarehouse, gmWarehouse))
    #ml_logger.info("upate staircases Object in floor/building done")

    #ml_logger.info("update corridor Object in floor/building ...")
    for corridor in floor._corridors:
        corridor.set_walls(update_walls(
            corridor._walls, wimWarehouse, wfmWarehouse, gmWarehouse))
    #ml_logger.info("update corridor Object in floor/building done")

    #ml_logger.info("convert floor to tensor ...")
    building_features, target = floor.to_tensor()  # 将建筑对象转换为张量
    #ml_logger.info("convert floor to tensor done")

    building_features = tf.cast(
        building_features, dtype=tf.float64)  # 将张量转换为浮点数
    target = tf.cast(target, dtype=tf.float64)  # 将张量转换为浮点数

    building_features = building_features.numpy()  # 将张量转换为numpy数组
    target = target.numpy()  # 将张量转换为numpy数组
    
    return building_features, target


# tensor to house
# 根据预测结果的对应的特征值生成一个新的对象HOUSE，用于转换为新的json文件或是转换输出到exl文件
@timing_decorator
def tensor_to_house(json_data, house_features, targets):

    # pdb.set_trace()

    i_house = House().json_to_house(json_data)
    house = copy.deepcopy(i_house)
    rf_len = Room().get_room_features_len()
    waf_len = Wall().get_wall_features_len()
    wif_len = Window().get_window_features_len()

    cursor = 0

    for room_index, room in enumerate(house._rooms):
        cursor = room_index * rf_len
        room_features = house_features[cursor:cursor+rf_len]

        for wall_index, wall in enumerate(room._walls):

            if wall._material is None:
                continue

            wall_features = room_features[4+wall_index *
                                          waf_len:4+(wall_index+1)*waf_len]

            window_features = wall_features[2+wimWarehouse.get_size():]

            wall._material = WallInsulation(area=wall._area, warehouse=wimWarehouse,
                                            material_type=wimWarehouse.get_material(np.argmax(wall_features[2:2+wimWarehouse.get_size()])))

            window = copy.deepcopy(wall.get_window())

            if window is not None:

                window_area = window_features[0]

                aspect_ratio = np.random.uniform(0.618, 1.618)
                window_width = np.sqrt(window_area * aspect_ratio)
                window_height = window_area / window_width

                if window_width > wall.get_width():
                    window_width = wall.get_width()
                    window_height = window_area / window_width
                if window_height > wall.get_height():
                    window_height = wall.get_height()
                    window_width = window_area / window_height

                wfa_ratio, wf_material = wfmWarehouse.get_best_wf(
                    window_width, window_height)
                # 生成一个窗户对象
                window = Window(area=window_area, width=window_width,
                                height=window_height, orientation=wall.get_orientation())
                window_frame = WindowFrame(
                    area=window_area*wfa_ratio, warehouse=wfmWarehouse, material_type=wf_material)
                # 生成一个窗框对象，将窗框添加窗户到上
                window.set_window_frame(window_frame)
                # 生成一个玻璃对象，将玻璃添加到窗户上
                glass = Glass(area=window_area*(1-wfa_ratio), warehouse=gmWarehouse,
                              material_type=gmWarehouse.get_random_material())
                window.set_glass(glass)
                wall.reset_window(window)


    return house


def convert_room_walls(room, room_features, waf_len, wimWarehouse, wfmWarehouse, gmWarehouse):

    # pdb.set_trace()

    for wall_index, wall in enumerate(room._walls):
        if wall._material is None:
            continue

        wall_features = room_features[4 + wall_index *
                                      waf_len:4 + (wall_index + 1) * waf_len]

        window_features = wall_features[2 + wimWarehouse.get_size():]
        wim_id = np.argmax(wall_features[2:2 + wimWarehouse.get_size()])
        material_type = wimWarehouse.get_material(wim_id)

        wall._material = WallInsulation(
            area=wall._area, warehouse=wimWarehouse, material_type=material_type)
        window = copy.deepcopy(wall.get_window())

        if window is not None:
            window_area = window_features[0]

            aspect_ratio = np.random.uniform(0.618, 1.618)
            window_width = np.sqrt(window_area * aspect_ratio)

            ml_logger.info(
                f"window_area {window_area} window_width {window_width}")
            window_height = window_area / window_width

            if window_width > wall.get_width():
                window_width = wall.get_width()
                window_height = window_area / window_width
            if window_height > wall.get_height():
                window_height = wall.get_height()
                window_width = window_area / window_height

            wfa_ratio, wf_material = wfmWarehouse.get_best_wf(
                window_width, window_height)

            window = Window(area=window_area, width=window_width,
                            height=window_height, orientation=wall.get_orientation())
            window_frame = WindowFrame(
                area=window_area * wfa_ratio, warehouse=wfmWarehouse, material_type=wf_material)

            window.set_window_frame(window_frame)
            glass = Glass(area=window_area * (1 - wfa_ratio), warehouse=gmWarehouse,
                          material_type=gmWarehouse.get_random_material())
            window.set_glass(glass)
            wall.reset_window(window)
        else:
            print("window is None")

    return room


def extract_building_features(building_features):

    # pdb.set_trace()

    # 判断是否为numpy数组，如果不是则转换为numpy数组
    # if not isinstance(building_features[0], np.ndarray):
    building_features = np.array(building_features[0])

    position_0 = 0
    position_1 = House.HOUSE_FEATURES * Floor.HOUSE_NUMBER
    position_2 = position_1 + PublicSpace.ROOM_FEATURES_LEN * Floor.STAIRCASE_NUMBER
    position_3 = position_2 + PublicSpace.ROOM_FEATURES_LEN * Floor.CORRIDORS_NUMBER

    # copy house_features from building_features
    house_features = building_features[position_0:position_1]
    # copy stair_features from building_features
    stair_features = building_features[position_1:position_2]
    # copy corridor_features from building_features
    corridor_features = building_features[position_2:position_3]

    return house_features, stair_features, corridor_features

@timing_decorator
def tensor_to_building(json_data, building_features, targets):

    #pdb.set_trace()

    i_building = Building().json_to_building(json_data)

    temp_doc_id = f"{i_building.get_doc_id()}_temp" 

    i_building.to_json(temp_doc_id)

    json_data = couchdb_pool.get_doc(temp_doc_id)

    print(json.dumps(json_data, indent=2, ensure_ascii=False))

    building = copy.deepcopy(i_building)

    hf_len = House.HOUSE_FEATURES
    rf_len = Room.ROOM_FEATURES_LEN
    waf_len = Wall.WALL_FEATURES_LEN
    wif_len = Window.WINDOW_FEATURES_LEN

    ml_logger.info(f"hf_len = {hf_len} rf_len = {rf_len} waf_len = {waf_len} wif_len = {waf_len} ")

    building_features = np.array(building_features)

    house_features, stair_features, corridor_features = extract_building_features(
        building_features)

    # position_0 = 0
    # position_1 = House.HOUSE_FEATURES*Floor.HOUSE_NUMBER
    # position_2 = position_1 + PublicSpace.ROOM_FEATURES_LEN*Floor.STAIRCASE_NUMBER
    # position_3 = position_2 + PublicSpace.ROOM_FEATURES_LEN*Floor.CORRIDORS_NUMBER

    # # copy house_features from building_features
    # house_features = building_features[position_0:position_1]
    # # copy stair_features from building_features
    # stair_features = building_features[position_1:position_2]
    # # copy corridor_features from building_features
    # corridor_features = building_features[position_2:position_3]

    #ml_logger.info(
    #    f"house_features shape = {house_features.shape} stair_features shape = {stair_features.shape} corridor_features shape = {corridor_features.shape} ")
    # pdb.set_trace()

    for floor_index, floor in enumerate(building._floors):
        # print(f"convert floor_index {floor_index}")
        # convert floor features
        for house_index, house in enumerate(floor._houses):

            h_features = house_features[house_index *
                                        hf_len:(house_index + 1) * hf_len]
            house = house.tensor_to_house(h_features, house_index)

            floor.update_house(copy.deepcopy(house), house_index)

        # convert stair features
        for stair_index, stair in enumerate(floor._staircases):
            cursor = stair_index * rf_len
            room_features = stair_features[cursor:cursor + rf_len]

            stair = stair.tensor_to_room(room_features)

            floor.update_staircase(copy.deepcopy(stair), stair_index)


        # convert corridor features
        for corridor_index, corridor in enumerate(floor._corridors):
            cursor = corridor_index * rf_len
            room_features = corridor_features[cursor:cursor + rf_len]

            corridor = corridor.tensor_to_room(room_features)

            floor.update_corridor(copy.deepcopy(corridor), corridor_index)

        # end for floor_index, floor in enumerate(building._floors)
    # end for i_building in buildings:
        building.update_floor(floor, floor_index)

    return building


def population_to_exl(x, y):

    x = np.vstack(x)
    y = np.vstack(y)
    result = np.concatenate((x, y), axis=1)
    # 将二维数组转换为 DataFrame
    df = pd.DataFrame(result)
    start_time = datetime.now()  # 设置开始时间
    time_format = start_time.strftime(
        "%Y%m%d-%H%M%S-%f")[:-3]  # 格式化时间，精确到100毫秒
    # 生成带时间序列的文件名
    file_name = f"./predicted_data/test_population/tp_{time_format}.xlsx"
    # 将 DataFrame 写入 Excel 文件
    df.to_excel(file_name, index=False)

    return file_name


def population_to_excel_T(x, y):
    # 将 x 和 y 转置，使数据按列存储

    x = np.vstack(x)
    y = np.vstack(y)
    result = np.concatenate((x, y), axis=1)
    result = result.T

    # 将二维数组转换为 DataFrame
    df = pd.DataFrame(result)

    start_time = datetime.now()  # 设置开始时间
    time_format = start_time.strftime(
        "%Y%m%d-%H%M%S-%f")[:-3]  # 格式化时间，精确到100毫秒
    # 生成带时间序列的文件名
    file_name = f"./predicted_data/test_population/tp_{time_format}.xlsx"
    # 将 DataFrame 转置后写入 Excel 文件
    df.to_excel(file_name, index=False, header=False)

    return file_name

# 把数据集导出到 Excel 表中


def export_to_excel(houses_x, houses_y, ds_size, path):
    print("\rExporting the training dataset to Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Dataset"

    rows_written = 0  # Keep track of the number of rows written
    rows_to_write = []  # List to accumulate rows

    for i, house_x in enumerate(houses_x):
        if i >= ds_size:
            break

        house_x = house_x.numpy()
        house_y = houses_y[i].numpy()
        row = list(house_x.flatten()) + list(house_y.flatten())
        rows_to_write.append(row)
        rows_written += 1

        if rows_written % 100 == 0:
            # Write the accumulated rows to the Excel file
            for row_to_write in rows_to_write:
                ws.append(row_to_write)
            rows_to_write = []  # Clear the list for the next batch
            sys.stdout.flush()
        sys.stdout.write("\rExporting to Excel: %d/%d" % (i+1, len(houses_x)))

    # Save any remaining rows
    if rows_to_write:
        for row_to_write in rows_to_write:
            ws.append(row_to_write)

    wb.save(path)
    print("\nDone.")


def import_from_excel(path):
    print("Importing the training dataset from Excel...")
    df = pd.read_excel(path)
    houses_x = []
    houses_y = []

    for i, row in df.iterrows():
        houses_x.append(row[0:House.HOUSE_FEATURES].tolist())
        houses_y.append(row[House.HOUSE_FEATURES:].tolist())
        sys.stdout.write("\rImporting from Excel: %d/%d" % (i+1, len(df)))
        sys.stdout.flush()

    print("\nDone.")

    return houses_x, houses_y


def generate_excel_filename(input_file=None):
    # 获取当前的日期和时间
    # now = datetime.datetime.now()
    now = datetime.now()

    # 将日期和时间转换为字符串
    timestamp = now.strftime('%Y%m%d%H%M%S')

    # 生成一个随机数
    random_num = random.randint(1000, 9999)

    # 创建文件名
    filename = f"{os.path.basename(input_file)}_{timestamp}_{random_num}.xlsx"

    # 创建文件路径
    filepath = os.path.join('./predicted_data/', filename)

    return filepath


def generate_cn_filename(file_path):
    try:
        # 获取文件名和扩展名
        file_name, file_ext = os.path.splitext(file_path)

        # 检查文件名中是否包含'_cn'
        if '_cn' in file_name:
            return file_path

        # 构建新的文件名
        new_file_name = file_name + "_cn" + file_ext

        return new_file_name
    except ValueError as e:
        print("文件名格式错误:", e)
        return None
    except Exception as e:
        print("发生未知错误:", e)
        return None


def test_tensor_to_building(json_doc_id):

    json_data = couchdb_pool.get_doc_json(json_doc_id)
    # print(json_data)
    # 随机生成一个 （1,259200）的矩阵
    building_feature = np.random.rand(1, 259200)

    building = tensor_to_building(json_data, building_feature, None)
    # print(building.to_json_cn(json_doc_id))


if __name__ == "__main__":

    test_tensor_to_building('16976193849716460')

    # x, y = generate_dataset(250)
    # export_to_excel(x, y, 250, "./datasets/training_dataset.xlsx")
    # import_from_excel("./datasets/training_dataset.xlsx")
    '''
    x,y = generate_test_population(50,'./houses_json/training_dataset_json/20230611-154118-562269/20230611_154221364406.json')
    population_to_exl(x,y)
    x = np.vstack(x)
    y = np.vstack(y)
    result = np.concatenate((x, y), axis=1)
    np.set_printoptions(threshold=np.inf)
    print("result ---", result)

    # 将二维数组转换为 DataFrame
    df = pd.DataFrame(result)

    # 将 DataFrame 写入 Excel 文件
    df.to_excel('output.xlsx', index=False)
    '''
