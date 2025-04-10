# Author: sunhy 2023.6.20

import openpyxl
import json
import codecs
import os
from openpyxl.styles import Border, Side, PatternFill, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import warnings
# 忽略字体问题警告
warnings.filterwarnings("ignore", category=UserWarning, message="cmap value too big/small")



class JSONToExcelConverter:
    def __init__(self, filter_keys):
        self.curr_row = 1
        self.filter_keys = filter_keys

    def write_titles(self, titles, sheet,title_height=30):

        title_font = Font(name='Arial', size=14, bold=True, color='000000')  # 创建字体样式

        for col, title in enumerate(titles, start=1):
            cell = sheet.cell(row=1, column=col, value=title)
            cell.fill = PatternFill(
                start_color="EB4221", end_color="EB4221", fill_type="solid")

            cell.font = title_font  # 应用标题字体样式

        # 设置标题行的高度
        sheet.row_dimensions[1].height = title_height


    def write_to_excel(self, data, sheet, row, col, curr_row):
        """
        Recursively writes JSON data to Excel sheet.
        """
        if isinstance(data, dict):
            for key, value in data.items():
                cell = sheet.cell(row=curr_row, column=col, value=str(key))
                if isinstance(value, (dict, list)):
                    curr_row = self.write_to_excel(
                        value, sheet, row+1, col+1, curr_row+1)
                else:
                    # value=str(value))
                    sheet.cell(row=curr_row, column=col+1, value=value)
                curr_row += 1
                # 设置键的底色
                cell.fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        elif isinstance(data, list):
            for item in data:
                curr_row = self.write_to_excel(
                    item, sheet, curr_row+1, col, curr_row)
        elif isinstance(data, str):
            sheet.cell(row=curr_row, column=col, value=str(data))
            curr_row += 1
        else:
            # value=str(data))
            sheet.cell(row=curr_row, column=col, value=data)
            curr_row += 1
        return curr_row

    def json_to_excel(self, filename=None, output_filename=None, data=None, model="FILE", sheet_name=None, titles=None):
        # def json_to_excel(self, filename, output_filename):
        """
        Converts JSON file to Excel file.
        """
        if model == "FILE":
            with codecs.open(filename, 'r', encoding='utf-8', errors='replace') as file:
                data = json.load(file)
        elif model == "DATA":
            data = data

        data = self.delete_multiple_properties(data, self.filter_keys)

        # wb = openpyxl.Workbook()
        # sheet = wb.active
        if not os.path.exists(output_filename):
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(output_filename)

        # wb = openpyxl.load_workbook(output_filename)

        if sheet_name is None:
            sheet = wb.create_sheet()
        else:
            sheet = wb.create_sheet(title=sheet_name)

        curr_row = 1

        if titles:
            self.write_titles(titles, sheet)
            curr_row += 1

        self.write_to_excel(data, sheet, 1, 1, curr_row)

        # 自适应列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # 冻结首行
        sheet.freeze_panes = 'A2'

        # 添加边框线条
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border

        wb.save(output_filename)

    def delete_multiple_properties(self, data, property_list):
        # 加载JSON数据
        # data = json.loads(json_data)

        # 递归函数用于处理嵌套对象
        def delete_property_recursive(obj, property_name):
            if isinstance(obj, dict):
                for key in list(obj.keys()):
                    if key in property_name:
                        del obj[key]
                    else:
                        delete_property_recursive(obj[key], property_name)
            elif isinstance(obj, list):
                for item in obj:
                    delete_property_recursive(item, property_name)

        # 调用递归函数删除属性
        delete_property_recursive(data, property_list)

        # 转换回JSON格式
        # updated_json = json.dumps(data)

        # return updated_json
        return data

    def compare_sheets(self, file_path, sheet1_name, sheet2_name):
        # 加载工作簿
        workbook = load_workbook(file_path)

        # 获取两个sheet
        sheet1 = workbook[sheet1_name]
        sheet2 = workbook[sheet2_name]

        # 创建红色填充样式
        red_fill = PatternFill(fill_type='solid', fgColor='FF0000')

        # 创建字体样式
        font_up = Font(name='Arial', size=10,
                       color='FF0000', bold=True)  # 红色向上箭头
        font_down = Font(name='Arial', size=10,
                         color='006400', bold=True)  # 绿色向下箭头

        # 遍历第二个sheet中的单元格
        for row in sheet2.iter_rows():
            for cell in row:
                # 判断单元格是否非空
                if cell.value is not None:
                    # 获取对应位置的第一个sheet中的单元格
                    compare_cell = sheet1[cell.coordinate]

                    # 比较两个单元格的值
                    if isinstance(cell.value, (int, float)) and isinstance(compare_cell.value, (int, float)):

                        if cell.value > compare_cell.value:

                            diff = round(cell.value - compare_cell.value, 4)
                            if compare_cell.value != 0:
                                diff_ratio = diff/compare_cell.value
                                cell.value = f"原值 {compare_cell.value} 现值 {cell.value} ↑ {diff}  变化率 + {diff_ratio:.2%}"
                            else:
                                cell.value = f"原值 {compare_cell.value} 现值 {cell.value} ↑ {diff}"

                            cell.font = font_up

                        elif cell.value < compare_cell.value:

                            diff = round(compare_cell.value - cell.value, 4)

                            if compare_cell.value != 0:
                                diff_ratio = diff/compare_cell.value
                                cell.value = f"原值 {compare_cell.value} 现值 {cell.value} ↓ {diff}  变化率 - {diff_ratio:.2%}"
                            else:
                                cell.value = f"原值 {compare_cell.value} 现值 {cell.value} ↓ {diff}"

                            cell.font = font_down

                    elif compare_cell.value != cell.value:
                        # 如果值不相同，将第二个sheet中的单元格底色设置为红色
                        cell.fill = red_fill

        # 删除默认的sheet
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)

        # 保存工作簿
        workbook.save(file_path)

    def insert_sheet_to_right(self, file_path, source_sheet_name, target_sheet_name):
        # 加载工作簿
        workbook = load_workbook(file_path)

        # 获取源sheet和目标sheet
        source_sheet = workbook[source_sheet_name]
        target_sheet = workbook[target_sheet_name]

        # 插入一个空列
        target_column = target_sheet.max_column + 1
        target_sheet.insert_cols(target_column)

        # 将源sheet的内容复制到目标sheet的右侧
        for row in range(1, source_sheet.max_row + 1):
            for col in range(1, source_sheet.max_column + 1):
                # 获取源sheet中的单元格
                source_cell = source_sheet.cell(row=row, column=col)

                # 获取目标sheet中对应位置的单元格
                target_cell = target_sheet.cell(row=row, column=target_column)

                # --------
                target_cell.value = source_cell.value

                # 复制字体样式
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    underline=source_cell.font.underline,
                    strike=source_cell.font.strike,
                    color=source_cell.font.color
                )
                # ------------
                # 复制源sheet单元格的值和样式到目标sheet
                # target_cell.value = source_cell.value
                # target_cell.font = Font(copy=source_cell.font)  # 复制字体样式
                # target_cell.border = Border(copy=source_cell.border)  # 复制边框样式
                if source_cell.border is not None:
                    target_cell.border = Border(
                        left=source_cell.border.left,
                        right=source_cell.border.right,
                        top=source_cell.border.top,
                        bottom=source_cell.border.bottom,
                        diagonal=source_cell.border.diagonal,
                        diagonal_direction=source_cell.border.diagonal_direction,
                        outline=source_cell.border.outline,
                        vertical=source_cell.border.vertical,
                        horizontal=source_cell.border.horizontal
                    )
                # target_cell.fill = PatternFill(copy=source_cell.fill)  # 复制填充样式
                if source_cell.fill is not None:
                    target_cell.fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        fgColor=source_cell.fill.fgColor,
                        bgColor=source_cell.fill.bgColor
                    )
                target_cell.number_format = source_cell.number_format
                target_cell.protection = Protection(locked=True)  # 锁定单元格

        # 调整目标sheet中锁定列的属性
        for row in target_sheet.iter_rows():
            target_cell = row[target_column - 1]
            target_cell.protection = Protection(locked=False)  # 解锁单元格

        # 保存工作簿
        workbook.save(file_path)

    def main(self, json_file, excel_file):
        self.json_to_excel(json_file, excel_file)


if __name__ == '__main__':
    converter = JSONToExcelConverter()
    json_file = input("请输入JSON文件路径：")
    excel_file = input("请输入Excel文件路径：")

    titles = input("请输入标题，以逗号分隔：").split(',')
    converter.json_to_excel(json_file, excel_file, titles=titles)
    # converter.main(json_file, excel_file,"test load json")
