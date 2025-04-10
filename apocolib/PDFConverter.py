# Author: sunhy
# Date: 2023-7-06
# Desc: 将 Excel 文件转换为 PDF 文件，支持多进程，支持汇总页，支持封面，支持水印，支持自定义页码起始值，支持加密 PDF 文件
# Usage:
#   excel_file = "./predicted_data/17023737560245870_20231219173254_8110.xlsx"
#   sheet_name = "Prediction result"
#   pdf_file = "./predicted_data/17023737560245870_20231219173254_8110.xlsx.pdf"
#   excel_to_pdf(excel_file, sheet_name, pdf_file)
#   excel_to_pdf(excel_file, sheet_name, pdf_file, summary_data=summary_data)
# 字体库，/usr/share/fonts/han_sans/SourceHanSansSC-VF.ttf, 用于支持中文，需要安装
# 依赖库：openpyxl, fpdf, tqdm, PyPDF2, concurrent.futures


import sys
sys.path.append("..")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import pdb
import json
import datetime
from apocolib.MlLogger import mlLogger as ml_logger
from tqdm import tqdm
import time
import math
import os
from PyPDF2 import PdfWriter,PdfReader,PdfMerger
from concurrent.futures import ProcessPoolExecutor, as_completed

class PDF(FPDF):

    def __init__(self,page_start_number=0):
        super().__init__()
        self.add_font('SourceHanSansSC-VF', '',
                      '/usr/share/fonts/han_sans/SourceHanSansSC-VF.ttf', uni=True)
        # 设置 PDF 页面尺寸和边距
        self.set_auto_page_break(auto=True, margin=15)
        self.set_font('SourceHanSansSC-VF', '', 10)

        # 设置起始页码
        self.page_start_number = page_start_number

    def header(self):
        # Logo
        self.image('./images/logo.png', 10, 8, 33)
        # Arial bold 15
        # self.set_font('Arial', 'B', 15)
        self.set_font('SourceHanSansSC-VF', '', 18)
        # Move to the right
        self.cell(100, 10)
        # Title
        self.cell(30, 10, '建筑空间设计和材料优化计算和预测报告', 0, 0, 'C')
        # Line break
        self.ln(20)
        # Draw black horizontal line
        # self.set_fill_color(0, 0, 0)  # Set fill color to black
        # self.rect(10, self.y, self.w - 20, 0.2, 'F')  # Draw line
    
    # 重写页码方法，用于灵活定义起始页码
    def page_no(self):
        return super().page_no() + self.page_start_number

    # Page footer
    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        # Arial italic 8
        # self.set_font('Arial', 'I', 8)
        # Page number
        self.cell(0, 10, '报告机构:  天启智源AI平台-APOCO DesignMate / 四川天启智源科技有限公司    第 ' +
                  str(self.page_no()) + ' 页', 0, 0, 'R')
        
    # 水印
    def watermark(self, text):

        # 将文本旋转并居中
        self.rotate(45)
        self.set_font('SourceHanSansSC-VF', '', 50)
        self.set_text_color(200, 200, 200)
        self.text(10, 160, text)
        self.text(30, 320, text)
        self.text(50, 480, text)
        # 重置文本旋转
        self.rotate(0)
        # 设置字体颜色
        self.set_text_color(0, 0, 0)
        # 重置字体和大小
        self.set_font('SourceHanSansSC-VF', '', 10)

def generate_pdf_page(iter_rows,temp_file,page_start_number=1,watermark="ai.apoco.com.cn"):

    pdf = PDF(page_start_number=page_start_number)
    pdf.add_page()

    #ml_logger.info(f"Generating PDF page {page_start_number}")
    # i 从 1 开始计算
    for i, row in enumerate(iter_rows, start=1):
        # 每30行一页，如何 i 是 30 的倍数，就新建一页
        if i % 30 == 0 : 
            pdf.add_page()
            pdf.watermark(watermark)

        row_is_empty = True
        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                #ml_logger.warning(f"cell value is None, cell: {cell}")
                continue
            pdf.cell(90, 8, str(cell_value), border=1)
            row_is_empty = False
        if not row_is_empty:
            pdf.ln()

    # 保存生成的PDF页
    pdf.output(temp_file)


def write_home_page( cover_title, project_name, organization, author, report_date,file_name="./predicted_data/temp/temp_page_0.pdf"):
    pdf = PDF()
    # 把footer 清除掉
    pdf.footer = lambda: None
    pdf.add_page()
    # 设置封面内容
    pdf.image("./images/bg.png", x=90, y=140, w=120, h=140)
    pdf.ln(65)
    pdf.set_font('SourceHanSansSC-VF', '', 18)
    pdf.cell(0, 10, '项目名称: ' + project_name, ln=True, align='C')
    pdf.ln(40)
    pdf.cell(0, 10, '报告单位: ' + organization, ln=True, align='L')
    pdf.cell(0, 10, '报告人: ' + author, ln=True, align='L')
    pdf.cell(0, 10, '报告时间: ' + report_date, ln=True, align='L')
    pdf.ln(20)
    pdf.set_font('SourceHanSansSC-VF', '', 8)

    home_page = file_name
    if(os.path.exists(home_page)):
        os.remove(home_page)

    pdf.output(home_page)
    #ml_logger.info(f"PDF {home_page} saved successfully!")
    return home_page


# 生成汇总页
def write_summary_page(summary_data,file_name="./predicted_data/temp/temp_page_999.pdf",page_start_number=999):

    if summary_data is None:
        return None
    
    table_data = [["材料型号", "面积", "价格", "金额"]]
    sorted_data = sorted(summary_data.items(),
                            key=lambda x: x[0])  # 按照 material 排序
    total_cost = 0.00
    for material, values in sorted_data:
        row = [str(material), round(float(values['area']), 4), round(
            float(values['price']), 2), round(float(values['cost']), 4)]
        total_cost += round(float(values['cost']), 4)
        table_data.append(row)
    table_data.append(['total cost', '/', '/', round(float(total_cost),4)])

    pdf = PDF(page_start_number=page_start_number)
    # 把footer 清除掉
    pdf.footer = lambda: None
    
    pdf.add_page()
    pdf.ln(20)
    pdf.set_font('SourceHanSansSC-VF', '', 18)
    pdf.cell(0, 10, '材料汇总数据', ln=True, align='L')
    pdf.ln(10)
    pdf.set_font('SourceHanSansSC-VF', '', 8)

    for i, row in enumerate(table_data):
        if i == 0:
            pdf.set_fill_color(255, 0, 0)
        else:
            pdf.set_fill_color(255, 255, 255)

        for j, data in enumerate(row):
            if j == 0:
                cell_len = 90
            else:
                cell_len = 30
            pdf.cell(cell_len, 8, str(data), border=1)
        pdf.ln()
    pdf.ln()



    pdf.output(file_name)

    return file_name

# 合并多个 PDF 文件
def merge_pdfs(output_file, temp_files):
    
    # 对文件名排序，避免顺序错乱
    temp_files.sort(key=lambda x: int(x.split('_')[-1].split('.')[0]))

    merger = PdfMerger()
    for temp_file in temp_files:
        #ml_logger.info(f"merging {temp_file}")
        merger.append(temp_file)
    
    merger.write(output_file)
    merger.close()

    #清理临时文件
    for temp_file in temp_files:
        if os.path.exists(temp_file):
            os.remove(temp_file)
            #ml_logger.info(f"remove temp file {temp_file}")

    return True

# pdf 加密
def encrypt_pdf(input_file, output_file, password="ai.apoco.com.cn"):

    pdf_writer = PdfWriter()
    pdf_reader = PdfReader(input_file)
    pdf_writer.append_pages_from_reader(pdf_reader)
    pdf_writer.encrypt(password)
    pdf_writer.write(output_file)
    return True
    
# 将 Excel 文件转换为 PDF 文件
def excel_to_pdf(excel_file, sheet_name, pdf_file, 
                 summary_data=None, cover_title='Report', project_name='Porject', 
                 organization='APOCO AI', author='APOCO AI', report_date=None, 
                 del_title=True,temp_path="./predicted_data/temp/",
                 watermark="ai.apoco.com.cn",password="ai.apoco.com.cn"):
    
    total_pages = 0

    if report_date is None:
        report_date = datetime.datetime.now().strftime("%Y-%m-%d")
    # 加载 Excel 文档
    workbook = load_workbook(excel_file)
    # 获取指定的 sheet

    sheet = workbook[sheet_name]
    if del_title:
        # 删除第一行标题
        sheet.delete_rows(1, 1)

    # 获取总行数和总页数
    total_rows = sheet.max_row
    page_rows = 30 # 35
    total_pages = total_rows // page_rows + 1
    temp_files = []
    # 生成封面
    temp_file = write_home_page(cover_title, project_name, organization, author, report_date,file_name=f"{temp_path}temp_page_0.pdf")
    if temp_file is not None:
        temp_files.append(temp_file) # 将封面的临时 PDF 加入列表

    # 多进程生成多个临时子文件，提高生成速度
    with ProcessPoolExecutor() as executor:

        page_batch = 300 # 生成 PDF 页的批次，每批次生成 300 页一个 PDF 文件

        p_id = math.ceil(total_pages / page_batch)  #
        futures = []

        for i in range(p_id):

            start_row = i * page_rows * page_batch + 1
            if start_row >= total_rows:
                return
            if (total_rows - start_row ) >= page_rows * page_batch:
                end_row =  start_row + (page_rows * page_batch)
            else:
                end_row = total_rows

            temp_rows = list(sheet.iter_rows(min_row=start_row, max_row=end_row)) # 避免迭代器被重复使用

            temp_file = f"{temp_path}temp_page_{i+1}.pdf"

            future = executor.submit(generate_pdf_page, temp_rows, temp_file, 
                                     page_start_number=i*page_batch ,#+1,
                                     watermark=watermark)
            
            #future.add_done_callback(lambda x: ml_logger.info(f"future done,p_id {i}") )
            futures.append(future)
            temp_files.append(temp_file)

        # 显示进度条，等待上述所有任务完成
        for future in tqdm(as_completed(futures), total=len(futures), desc="Generating PDF pages"):
            # 获取任务结果
            result = future.result() 

    # 生成汇总页
    temp_file = write_summary_page(summary_data,file_name=f"{temp_path}temp_page_{p_id+1}.pdf",
                                   page_start_number=p_id*page_batch+1)
    if temp_file is not None:
        temp_files.append(temp_file)
    
    # 将home_page ,pdfs ,summary page 合并 1个 PDF 文件
    merge_pdfs(pdf_file,temp_files)

    # 加密 PDF 文件
    ml_logger.info("Encrypting PDF file...")
    encrypt_pdf(pdf_file, pdf_file, password=password)
   
    ml_logger.info("PDF file saved successfully!")

    return total_pages

if __name__ == "__main__":
    # 示例用法
    excel_file = "./predicted_data/17023737560245870_20231219173254_8110.xlsx"
    sheet_name = "Prediction result"
    pdf_file = "./predicted_data/17023737560245870_20231219173254_8110.xlsx.pdf"

    excel_to_pdf(excel_file, sheet_name, pdf_file)
