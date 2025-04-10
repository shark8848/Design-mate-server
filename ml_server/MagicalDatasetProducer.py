import numpy as np
import dataSetBaseParamters as ds_bp
from HouseDataGenerator import HouseDataGenerator,CustomJSONEncoder
import pandas as pd
import json
from datetime import datetime
import sys
import pdb
import copy
import random
import os
from openpyxl import Workbook, load_workbook

sys.path.append("..")
from apocolib.MlLogger import mlLogger as ml_logger
from apocolib.RabbitMQProducer import RabbitMQProducer
from apocolib.MQueueManager import MQueueManager

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
#from websocket_server.webSocketServer import external_push_message as expm

#rmqp = RabbitMQProducer(queue_manager=MQueueManager())

# 生成数据集
def generate_dataset(size=None):

    houses_x = []
    houses_y = []

    ds_size = size if size is not None else ds_bp.get_dataset_size()
    data_dir = f"training_dataset_json/{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}"
    for i in range(ds_size):

        house_generator = HouseDataGenerator(ds_bp.get_random_rooms())
        house = house_generator.generate_house_data(data_dir)
        house_features,house_targets = house_generator.ConvertHouseToArray(house)

        houses_x.append(house_features)
        houses_y.append(house_targets)
        sys.stdout.write("\rGenerate dataset: %d/%d" % (i+1, ds_size))
        sys.stdout.flush()
        #time.sleep(0.1)
        #expm("AC2NNetTrainer",f"Generate dataset:{i+1}/{ds_size}")
        #rmqp.publish(f"Generate dataset:{i+1}/{ds_size}")
        #rmqp.close()
    return houses_x, houses_y

#_ds_X ,_ds_Y = generate_dataset()
_ds_X = []
_ds_Y = [] #generate_dataset()


def init_dataset(fixed=1):
    if fixed == 1:
        return _ds_X, _ds_Y
    else:
        return generate_dataset()


def generate_dataset_columns():
    max_num_rooms = ds_bp.get_max_num_rooms()
    max_num_walls = ds_bp.get_max_num_walls()
    max_num_windows = ds_bp.get_max_num_windows()
    wall_material_len = ds_bp.get_WIM_num()
    glass_type_len = ds_bp.get_GM_num()
    wf_material_len = ds_bp.get_WF_num() # sunhy 2023.6.4

    feature_columns = []
    target_columns = []

    for room_idx in range(max_num_rooms):
        feature_columns.extend([
            f'room_{room_idx}_type',
            f'room_{room_idx}_area',
            f'room_{room_idx}_total_wall_area',
            f'room_{room_idx}_total_window_area',
        ])
        for wall_idx in range(max_num_walls):
            feature_columns.extend([
                f'room_{room_idx}_wall_{wall_idx}_area',
                f'room_{room_idx}_wall_{wall_idx}_orientation',
            ])
        for window_idx in range(max_num_windows):
            feature_columns.extend([
                f'room_{room_idx}_window_{window_idx}_area',
                f'room_{room_idx}_window_{window_idx}_orientation',
            ])
        feature_columns.extend([f'room_{room_idx}_wall_material_{i}' for i in range(wall_material_len)])
        feature_columns.extend([f'room_{room_idx}_glass_type_{i}' for i in range(glass_type_len)])
        feature_columns.extend([f'room_{room_idx}_wf_material_{i}' for i in range(wf_material_len)]) # sunhy 2023.6.4

        target_columns.extend([
            f'room_{room_idx}_total_cost',
            f'room_{room_idx}_total_avg_k',
        ])

    # 在 target_columns 中添加总的 cost 和 avg_k
    target_columns.extend([
        'h_total_cost',
        'h_total_avg_k',
    ])

    return feature_columns, target_columns

def export_to_excel(houses_x, houses_y,feature_columns, target_columns, filename):

 # 检查 feature_columns 和 houses_x 的长度是否匹配
    if len(feature_columns) != len(houses_x[0]):
        print(f"Length of feature_columns: {len(feature_columns)}")
        print(f"Length of houses_x data: {len(houses_x[0])}")
        raise ValueError("The length of feature_columns and houses_x data do not match.")
    
    # 检查 target_columns 和 houses_y 的长度是否匹配
    if len(target_columns) != len(houses_y[0]):
        print(f"Length of target_columns: {len(target_columns)}")
        print(f"Length of houses_y data: {len(houses_y[0])}")
        raise ValueError("The length of target_columns and houses_y data do not match.")

    # 将houses_x和houses_y转换为pandas DataFrame
    houses_x_df = pd.DataFrame(houses_x, columns=feature_columns)
    houses_y_df = pd.DataFrame(houses_y, columns=target_columns)

    # 将houses_x_df和houses_y_df合并为一个DataFrame
    combined_df = pd.concat([houses_x_df, houses_y_df], axis=1)

    # 将DataFrame导出到Excel文件
    combined_df.to_excel(filename, index=False)

# 生成数据集 test
#houses_x, houses_y  = generate_dataset(size=2)
#feature_columns, target_columns = generate_dataset_columns()
# 导出到Excel文件
#export_to_excel(houses_x, houses_y, feature_columns, target_columns,'./datasets/houses_dataset.xlsx')
# 将个体转换为参数, 入口参数 individual = generate_test_house()

def print_individual(individual):
    ml_logger.info("Individual:", individual)
    house = individual

    max_num_rooms = ds_bp.get_max_num_rooms()
    max_num_walls = ds_bp.get_max_num_walls()
    max_num_windows = ds_bp.get_max_num_windows()
    wall_material_len = ds_bp.get_WIM_num()
    glass_type_len = ds_bp.get_GM_num()
    wf_material_len = ds_bp.get_WF_num() # sunhy 2023.6.4

    room_feature_len = (
        4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len + wf_material_len
    )

    for room_index in range(max_num_rooms):
        ml_logger.info(f"Room {room_index + 1}:")

        room_start = room_index * room_feature_len

        room_type = house[room_start]
        room_area = house[room_start + 1]
        if room_area == 0:
            ml_logger.info("  - No data for this room")
            continue

        ml_logger.info(f"  - Room type: {room_type}")
        ml_logger.info(f"  - Room area: {room_area:.2f}")
        ml_logger.info(f"  - Total wall area: {house[room_start + 2]:.2f}")
        ml_logger.info(f"  - Total window area: {house[room_start + 3]:.2f}")

        ml_logger.info("  - Wall areas and orientations:")
        for i in range(max_num_walls):
            area = house[room_start + 4 + i]
            orientation = house[room_start + 4 + max_num_walls + i]
            if area > 0:
                ml_logger.info(f"    - Area: {area:.2f}, Orientation: {orientation}")

        ml_logger.info("  - Window areas and orientations:")
        for i in range(max_num_windows):
            area = house[room_start + 4 + 2 * max_num_walls + i]
            orientation = house[room_start + 4 + 2 *
                                max_num_walls + max_num_windows + i]
            if area > 0:
                ml_logger.info(f"    - Area: {area:.2f}, Orientation: {orientation}")

        ml_logger.info("  - Wall material:")
        for i in range(wall_material_len):
            value = house[room_start + 4 + 2 *
                          max_num_walls + 2 * max_num_windows + i]
            if value > 0:
                ml_logger.info(f"    - Material {i + 1}: {value}")

        ml_logger.info("  - Glass type:")
        for i in range(glass_type_len):
            value = house[room_start + 4 + 2 * max_num_walls +
                          2 * max_num_windows + wall_material_len + i]
            if value > 0:
                ml_logger.info(f"    - Glass type {i + 1}: {value}")

        #--- 6.4 begin
        ml_logger.info("  - WF material:")
        for i in range(wf_material_len):
            value = house[room_start + 4 + 2 * max_num_walls +
                          2 * max_num_windows + wall_material_len + glass_type_len + i]
            if value > 0:
                ml_logger.info(f"    - WF material {i + 1}: {value}")
        #--- 6.4 end

        ml_logger.info("")

# test print
#houses_x, houses_y, feature_columns, target_columns = generate_dataset()
#print_individual(houses_x[0])

def print_individual_to_exl(individual, output_col, excel_file="./predicted_data/output.xlsx"):
    # Original function code
    house = individual
    max_num_rooms = ds_bp.get_max_num_rooms()
    max_num_walls = ds_bp.get_max_num_walls()
    max_num_windows = ds_bp.get_max_num_windows()
    wall_material_len = ds_bp.get_WIM_num()
    glass_type_len = ds_bp.get_GM_num()
    wf_material_len = ds_bp.get_WF_num() # sunhy 2023.6.4
    
    room_feature_len = (
        4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len + wf_material_len
    )
    # Load or create a workbook
    try:
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        wb = Workbook()

    # Select or create a worksheet
    ws = wb.active
    ws.title = "Individuals"

    # Write data to specified column
    row = 1
    ws.cell(row=row, column=output_col, value="Individual:")
    ws.cell(row=row, column=output_col + 1, value=str(individual))
    row += 1

    for room_index in range(max_num_rooms):
        ws.cell(row=row, column=output_col, value=f"Room {room_index + 1}:")
        row += 1

        room_start = room_index * room_feature_len

        room_type = house[room_start]
        room_area = house[room_start + 1]
        if room_area == 0:
            ws.cell(row=row, column=output_col, value="  - No data for this room")
            row += 1
            continue

        ws.cell(row=row, column=output_col, value=f"  - Room type: {room_type}")
        row += 1
        ws.cell(row=row, column=output_col, value=f"  - Room area: {room_area:.2f}")
        row += 1
        ws.cell(row=row, column=output_col, value=f"  - Total wall area: {house[room_start + 2]:.2f}")
        row += 1
        ws.cell(row=row, column=output_col, value=f"  - Total window area: {house[room_start + 3]:.2f}")
        row += 1

        ws.cell(row=row, column=output_col, value="  - Wall areas and orientations:")
        row += 1
        for i in range(max_num_walls):
            area = house[room_start + 4 + i]
            orientation = house[room_start + 4 + max_num_walls + i]
            if area > 0:
                ws.cell(row=row, column=output_col, value=f"    - Area: {area:.2f}, Orientation: {orientation}")
                row += 1

        ws.cell(row=row, column=output_col, value="  - Window areas and orientations:")
        row += 1
        for i in range(max_num_windows):
            area = house[room_start + 4 + 2 * max_num_walls + i]
            orientation = house[room_start + 4 + 2 * max_num_walls + max_num_windows + i]
            if area > 0:
                ws.cell(row=row, column=output_col, value=f"    - Area: {area:.2f}, Orientation: {orientation}")
                row += 1

        ws.cell(row=row, column=output_col, value="  - Wall material:")
        row += 1
        for i in range(wall_material_len):
            value = house[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + i]
            if value > 0:
                ws.cell(row=row, column=output_col, value=f"    - Material {i + 1}: {value}")
                row += 1

        ws.cell(row=row, column=output_col, value="  - Glass type:")
        row += 1
        for i in range(glass_type_len):
            value = house[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + i]
            if value > 0:
                ws.cell(row=row, column=output_col, value=f"    - Glass type {i + 1}: {value}")
                row += 1

        #-----begin sunhy 2023.06.4
        ws.cell(row=row, column=output_col, value="  - WF material:")
        row += 1
        for i in range(wf_material_len):
            value = house[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len + i]
            if value > 0:
                ws.cell(row=row, column=output_col, value=f"    - WF material {i + 1}: {value}")
                row += 1

        #----- end sunhy 2023.06.4
        row += 1

    # Save the workbook
    wb.save(excel_file)

#------------- begin 6.5
#from openpyxl import load_workbook, Workbook
#from openpyxl.utils import get_column_letter


def print_individual_to_exl_v2(individual, output_col, excel_file="./predicted_data/output.xlsx"):
    house = individual
    max_num_rooms = ds_bp.get_max_num_rooms()
    max_num_walls = ds_bp.get_max_num_walls()
    max_num_windows = ds_bp.get_max_num_windows()
    wall_material_len = ds_bp.get_WIM_num()
    glass_type_len = ds_bp.get_GM_num()

    room_feature_len = (
        4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len
    )

    wb = load_workbook(excel_file) if os.path.exists(excel_file) else Workbook()
    ws = wb.active
    ws.title = "Individuals"

    write_tree_structure(ws, house, output_col, max_num_rooms, max_num_walls, max_num_windows,
                         wall_material_len, glass_type_len, room_feature_len)

    wb.save(excel_file)


def write_tree_structure(ws, house, output_col, max_num_rooms, max_num_walls, max_num_windows,
                         wall_material_len, glass_type_len, room_feature_len, row=1, indent_level=0):

    def write_cell(row, column, value):
        column_letter = get_column_letter(column)
        ws.cell(row=row, column=column, value=value)

    indent = "  " * indent_level

    write_cell(row, output_col, "Individual:")
    write_cell(row, output_col + 1, str(house))
    row += 1

    for room_index in range(max_num_rooms):
        room_start = room_index * room_feature_len

        room_type = house[room_start]
        room_area = house[room_start + 1]

        write_cell(row, output_col, f"{indent}Room {room_index + 1}:")
        row += 1

        if room_area == 0:
            write_cell(row, output_col, f"{indent}  - No data for this room")
            row += 1
            continue

        write_cell(row, output_col, f"{indent}  - Room type: {room_type}")
        row += 1
        write_cell(row, output_col, f"{indent}  - Room area: {room_area:.2f}")
        row += 1
        write_cell(row, output_col, f"{indent}  - Total wall area: {house[room_start + 2]:.2f}")
        row += 1
        write_cell(row, output_col, f"{indent}  - Total window area: {house[room_start + 3]:.2f}")
        row += 1

        write_cell(row, output_col, f"{indent}  - Wall areas and orientations:")
        row += 1
        for i in range(max_num_walls):
            area = house[room_start + 4 + i]
            orientation = house[room_start + 4 + max_num_walls + i]
            if area > 0:
                write_cell(row, output_col, f"{indent}    - Area: {area:.2f}, Orientation: {orientation}")
                row += 1

        write_cell(row, output_col, f"{indent}  - Window areas and orientations:")
        row += 1
        for i in range(max_num_windows):
            area = house[room_start + 4 + 2 * max_num_walls + i]
            orientation = house[room_start + 4 + 2 * max_num_walls + max_num_windows + i]
            if area > 0:
                write_cell(row, output_col, f"{indent}    - Area: {area:.2f}, Orientation: {orientation}")
                row += 1

        write_cell(row, output_col, f"{indent}  - Wall material:")
        row += 1
        for i in range(wall_material_len):
            value = house[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + i]
            if value > 0:
                write_cell(row, output_col, f"{indent}    - Material {i + 1}: {value}")
                row += 1

        write_cell(row, output_col, f"{indent}  - Glass type:")
        row += 1
        for i in range(glass_type_len):
            value = house[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + i]
            if value > 0:
                write_cell(row, output_col, f"{indent}    - Glass type {i + 1}: {value}")
                row += 1

        row += 1
        write_tree_structure(ws, house, output_col, max_num_rooms, max_num_walls, max_num_windows,
                             wall_material_len, glass_type_len, room_feature_len, row, indent_level + 1)

#------------- end 6.5

def generate_excel_filename(input_file=None):
    # 获取当前的日期和时间
    #now = datetime.datetime.now()
    now = datetime.now()

    # 将日期和时间转换为字符串
    timestamp = now.strftime('%Y%m%d%H%M%S')

    # 生成一个随机数
    random_num = random.randint(1000, 9999)

    # 创建文件名
    filename = f"{os.path.basename(input_file)}_{timestamp}_{random_num}.xlsx"

    # 创建文件路径
    filepath = os.path.join('./predicted_data/', filename)

    return filepath

def generate_individual(house, wall_material_len, glass_type_len, wf_material_len, max_num_rooms, max_num_walls, max_num_windows): # sunhy 2023.6.4 add arg:'wf_material_len'
#def generate_individual(house, wall_material_len, glass_type_len, max_num_rooms, max_num_walls, max_num_windows):
    new_house = copy.deepcopy(house)

    # 遍历每个房间
    for room_index, room in enumerate(new_house["rooms"]):
        total_window_area = 0
        # 遍历每个墙壁
        for wall_index, wall in enumerate(room['walls']):

            # 随机窗户面积,限制调整幅度(-15% ~ +15%)和上下限值

            if wall['window'] is not None:
                original_window_area = wall['window']['area']
                min_area = max(0.85 * original_window_area, 0)
                max_area = min(1.15 * original_window_area, wall['area'])
                window_area = random.uniform(min_area, max_area)
                total_window_area += window_area
                wall['window']['area'] = window_area

            '''
            if wall['window'] is not None:
                window_area = random.uniform(0, wall['area'])
                total_window_area += window_area
                wall['window']['area'] = window_area
                '''

        # 更新房间中的总窗户面积
        room['total_window_area'] = total_window_area

        # 随机选择墙壁保温材料
        wall_material = random.randint(0, wall_material_len - 1)
        for wall in room['walls']:
            wall['insulation_material']['key'] = wall_material

        # 随机选择窗户玻璃材料
        glass_material = random.randint(0, glass_type_len - 1)
        for wall in room['walls']:
            if wall['window'] is not None:
                wall['window']['glass_material']['key'] = glass_material

        # 随机选择窗框材料 added by sunhy ,6.4
        wf_material = random.randint(0, wf_material_len - 1)
        for wall in room['walls']:
            if wall['window'] is not None:
                wall['window']['wf_material']['key'] = wf_material
        


    house_features = house_to_features(new_house, max_num_rooms, max_num_walls, max_num_windows, wall_material_len, glass_type_len , wf_material_len ) # added arg:wf_material_len,sunhy 6.4
    return house_features

# 为墙壁保温材料创建一个独热编码向量
def one_hot_wall_material(wall_material_key, wall_material_len):
    wall_material_encoding = [0] * wall_material_len
    wall_material_encoding[wall_material_key] = 1
    return wall_material_encoding


# 为窗户玻璃材料创建一个独热编码向量
def one_hot_glass_material(glass_material_key, glass_type_len):
    glass_material_encoding = [0] * glass_type_len
    glass_material_encoding[glass_material_key] = 1
    return glass_material_encoding


# 为窗框材料创建一个独热编码向量 added by sunhy. 2023.6.4
def one_hot_wf_material(wf_material_key, wf_material_len):
    wf_material_encoding = [0] * wf_material_len
    wf_material_encoding[wf_material_key] = 1
    return wf_material_encoding

def house_to_features(house, max_num_rooms, max_num_walls, max_num_windows, wall_material_len, glass_type_len, wf_material_len):

    room_feature_len = 4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len + wf_material_len
    house_features = [0] * max_num_rooms * room_feature_len

    for room_index, room in enumerate(house["rooms"]):
        room_start = room_index * room_feature_len

        # 添加房间特征
        house_features[room_start] = room['room_type']
        house_features[room_start + 1] = room['room_area']
        house_features[room_start + 2] = room['total_wall_area']
        house_features[room_start + 3] = room['total_window_area']

        for wall_index, wall in enumerate(room['walls']):
            house_features[room_start + 4 + wall_index] = wall['area']
            house_features[room_start + 4 + max_num_walls + wall_index] = wall['orientation']

            if wall['window'] is not None:
                window_index = wall['window']['orientation']
                house_features[room_start + 4 + 2 * max_num_walls + window_index] = wall['window']['area']
                house_features[room_start + 4 + 2 * max_num_walls + max_num_windows + window_index] = wall['window']['orientation']

        #pdb.set_trace()
        # 添加墙壁保温材料特征
        wall_material_encoding = [0] * wall_material_len
        if room['walls']:
            wall_material_encoding = one_hot_wall_material(room['walls'][0]['insulation_material']['key'], wall_material_len)

        house_features[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows: room_start +
               4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len] = wall_material_encoding

        # 添加窗户玻璃材料特征
        glass_material_encoding = [0] * glass_type_len
        if room['walls'] and room['walls'][0]['window']:
            glass_material_encoding = one_hot_glass_material(room['walls'][0]['window']['glass_material']['key'], glass_type_len)
            #house_features[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len] = glass_material_encoding
        house_features[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len: room_start +
                   4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len] = glass_material_encoding


        # 添加窗框材料特征 added by sunhy 2023.06.4
        wf_material_encoding = [0] * wf_material_len
        if room['walls'] and room['walls'][0]['window']:
            wf_material_encoding = one_hot_wf_material(room['walls'][0]['window']['wf_material']['key'], wf_material_len)
        house_features[room_start + 4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len: room_start +
                   4 + 2 * max_num_walls + 2 * max_num_windows + wall_material_len + glass_type_len + wf_material_len] = wf_material_encoding

    return house_features


def generate_population(base_house, num_individuals, wall_material_len, glass_type_len, wf_material_len, max_num_rooms, max_num_walls, max_num_windows): # added arg:wf_material_len,sunhy 2023.6.4
    population = []

    for i in range(num_individuals):
        individual = generate_individual(base_house, wall_material_len, glass_type_len, wf_material_len, max_num_rooms, max_num_walls, max_num_windows) # added arg:wf_material_len,sunhy 2023.6.4
        population.append(individual)

    return population

# 当loadFromFile=None 随机生成,否则从指定json文件读取
def generate_test_population(num_individuals=50,loadFromFile=None):
    base_house = None
    if loadFromFile is None:
        house_generator = HouseDataGenerator(ds_bp.get_random_rooms())
        data_dir = f"population_dataset_json/{datetime.now().strftime('%Y%m%d')}"
        base_house = house_generator.generate_house_data(data_dir)
        #base_house = generate_house_data(ds_bp.get_random_rooms())
    else:
        with open(loadFromFile, 'r') as f:
            base_house = json.load(f)

    num_individuals = num_individuals
    wall_material_len = ds_bp.get_WIM_num()

    glass_type_len = ds_bp.get_GM_num()
    max_num_rooms = ds_bp.get_max_num_rooms()
    max_num_walls = ds_bp.get_max_num_walls()
    max_num_windows = ds_bp.get_max_num_windows()
    wf_material_len = ds_bp.get_WF_num()

    population = generate_population(base_house, num_individuals, wall_material_len, glass_type_len, wf_material_len, max_num_rooms, max_num_walls, max_num_windows) # added arg:wf_material_len,sunhy 2023.6.4
    return population

#print(generate_test_population(10))
