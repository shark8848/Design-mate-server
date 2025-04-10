import numpy as np
import dataSetBaseParamters as ds_bp
from HouseDataGenerator import HouseDataGenerator,CustomJSONEncoder
import pandas as pd
import json
from datetime import datetime
import sys
import pdb
import copy
import random
import os
from openpyxl import Workbook, load_workbook

sys.path.append("..")
from apocolib.MlLogger import mlLogger as ml_logger
from apocolib.RabbitMQProducer import RabbitMQProducer
from apocolib.MQueueManager import MQueueManager

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from BuildingSpaceBase import *
from MaterialWarehouse import GlassMaterialWarehouse,WallInsulationMaterialWarehouse,WindowFrameMaterialWarehouse
#from websocket_server.webSocketServer import external_push_message as expm

#rmqp = RabbitMQProducer(queue_manager=MQueueManager())
gmWarehouse = GlassMaterialWarehouse()
wimWarehouse = WallInsulationMaterialWarehouse()
wfmWarehouse = WindowFrameMaterialWarehouse()
_ss = SpaceStandard() #创建空间标准对象

def check_data_type(tensor):
    if tf.is_tensor(tensor):
        data_type = tensor.dtype
        if data_type.is_floating:
            #print("张量中的数据类型为浮点型")
            return "float"
        elif data_type.is_integer:
            #print("张量中的数据类型为整型")
            return "int"
        else:
            #print("张量中的数据类型为对象")
            return "object"
    else:
        #print("传递的数据不是张量")
        return "error tensor"

# 生成数据集
def generate_dataset(size=1000,model='create',dataset_file='./datasets/training_dataset.xlsx'): # create;reload

    houses_x = []
    houses_y = []

    ds_size = size #size if size is not None else ds_bp.get_dataset_size()
    print("Preparing to generate the training dataset. It may take a few minutes. Please be patient.")
    if model == 'create':
        houses_creator = HousesCreator(ds_size)
            # 生成一个房屋对象
        houses = houses_creator.make_houses()

        data_dir = f"./houses_json/training_dataset_json/{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}"

        for i,house in enumerate(houses):

            file_serias = datetime.now().strftime("%Y%m%d_%H%M%S%f")[:21]
            house.save_to_json(f"{data_dir}/{file_serias}.json")
            house.save_to_json_cn(f"{data_dir}/{file_serias}_cn.json")
            houses_features ,target = house.to_tensor()
            houses_features = tf.cast(houses_features, dtype=tf.float64)
            target = tf.cast(target, dtype=tf.float64)

            #print(" houses_features is tensor ",tf.is_tensor(houses_features), "houses_features shape ",houses_features.shape, "data_type ",check_data_type(houses_features))
            #print(" target is tensor ",tf.is_tensor(target), "target shape ",target.shape, "data_type ",check_data_type(target))


            houses_x.append(houses_features)
            houses_y.append(target)
            sys.stdout.write("\rGenerate dataset: %d/%d houses_features shape: %s data_type is %s target shape %s  data_type is %s " % \
                    (i+1, ds_size,houses_features.shape, check_data_type(houses_features),target.shape,check_data_type(target)))
            sys.stdout.flush()
            #time.sleep(0.1)
            #expm("AC2NNetTrainer",f"Generate dataset:{i+1}/{ds_size}")
            #rmqp.publish(f"Generate dataset:{i+1}/{ds_size}")
            #rmqp.close()
        export_to_excel(houses_x, houses_y, size, dataset_file)
        return houses_x, houses_y

    elif model == 'reload':
        houses_x, houses_y = import_from_excel("./datasets/training_dataset.xlsx")
        return houses_x, houses_y
    else:
        print('Invalid create dataset model {model}')
        return [],[]

def generate_test_population(num_individuals=50,loadFromFile=None):

    json_data = None
    with open(loadFromFile, 'r') as f:
        json_data = json.load(f)

    houses_x = []
    houses_y = []

    ds_size = num_individuals #size if size is not None else ds_bp.get_dataset_size()

    i_house = House().json_to_house(json_data)

    for i in range(num_individuals):
        house = copy.deepcopy(i_house)

        # 随机对house 对象中的窗户面积进行修改
        #pdb.set_trace()
        for room in house._rooms:
            for wall in room._walls:

                wall._material = WallInsulation( area = wall._area, warehouse = wimWarehouse, 
                                    material_type = wimWarehouse.get_random_material() )
                window = copy.deepcopy(wall.get_window())

                if window is not None:

                    window_area = min(random.uniform(0.85, 1.15) * window._area, wall._area)
                    aspect_ratio = np.random.uniform(0.618, 1.618)
                    window_width = np.sqrt(window_area * aspect_ratio)
                    window_height = window_area / window_width

                    if window_width > wall.get_width():
                        window_width = wall.get_width()
                        window_height = window_area / window_width
                    if window_height > wall.get_height():
                        window_height = wall.get_height()
                        window_width = window_area / window_height

                    wfa_ratio, wf_material = wfmWarehouse.get_best_wf(window_width, window_height)
                    # 生成一个窗户对象
                    window = Window(area=window_area, width=window_width,height=window_height,orientation=wall.get_orientation())
                    window_frame = WindowFrame(area=window_area*wfa_ratio,warehouse = wfmWarehouse,material_type=wf_material)
                    # 生成一个窗框对象，将窗框添加窗户到上
                    window.set_window_frame(window_frame)
                    # 生成一个玻璃对象，将玻璃添加到窗户上
                    glass = Glass(area=window_area*(1-wfa_ratio),warehouse = gmWarehouse,material_type=gmWarehouse.get_random_material())
                    window.set_glass(glass)
                    wall.reset_window(window)


        house_features ,target = house.to_tensor()
        house_features = tf.cast(house_features, dtype=tf.float64)
        target = tf.cast(target, dtype=tf.float64)

        house_features = house_features.numpy()
        target = target.numpy()


        sys.stdout.write("\rGenerate test population: %d/%d houses_features shape: %s data_type is %s target shape %s  data_type is %s" % \
                (i+1, num_individuals,house_features.shape, check_data_type(house_features),target.shape,check_data_type(target)))
        sys.stdout.flush()

        houses_x.append(house_features)
        houses_y.append(target)

    population_to_exl(houses_x,houses_y)

    return houses_x, houses_y

# tensor to house
# 根据预测结果的对应的特征值生成一个新的对象HOUSE，用于转换为新的json文件或是转换输出到exl文件
def tensor_to_house(json_data,house_features,targets):

    #pdb.set_trace()

    i_house = House().json_to_house(json_data)
    house = copy.deepcopy(i_house)
    rf_len = Room().get_room_features_len()
    waf_len = Wall().get_wall_features_len()
    wif_len = Window().get_window_features_len()

    cursor = 0

    for room_index,room in enumerate(house._rooms):
        cursor = room_index * rf_len
        room_features = house_features[cursor:cursor+rf_len]
        #room._type = room_features[0]
        #room._area = room_features[1]
        print("-------------------------------------------------------------------------------")
        print(f"room[{room_index}] room_features {room_features}")

        for wall_index,wall in enumerate(room._walls):

            if wall._material is None:
                continue

            wall_features = room_features[4+wall_index*waf_len:4+(wall_index+1)*waf_len]
            print("-------------------------------------------------------------------------------")
            print(f"room[{room_index}] wall[{wall_index}] wall_features {wall_features}") 
            #window_features = wall_features[2:2+wif_len]
            window_features = wall_features[2+wimWarehouse.get_size():] #2+wif_len]
            print("-------------------------------------------------------------------------------")
            print(f"room[{room_index}] wall[{wall_index}] window_features {window_features}") 

            #wall._area = wall_features[0]
            #wall._orientation = wall_features[1].astype(int)
            #print("wall._area ",wall._area," wall._orientation ",wall._orientation)

            wall._material = WallInsulation( area = wall._area, warehouse = wimWarehouse,
                                material_type = wimWarehouse.get_material(np.argmax(wall_features[2:2+wimWarehouse.get_size()])) )

            window = copy.deepcopy(wall.get_window())

            if window is not None:
                #-----------
                window_area = window_features[0]

                aspect_ratio = np.random.uniform(0.618, 1.618)
                window_width = np.sqrt(window_area * aspect_ratio)
                window_height = window_area / window_width

                if window_width > wall.get_width():
                    window_width = wall.get_width()
                    window_height = window_area / window_width
                if window_height > wall.get_height():
                    window_height = wall.get_height()
                    window_width = window_area / window_height

                wfa_ratio, wf_material = wfmWarehouse.get_best_wf(window_width, window_height)
                # 生成一个窗户对象
                window = Window(area=window_area, width=window_width,height=window_height,orientation=wall.get_orientation())
                window_frame = WindowFrame(area=window_area*wfa_ratio,warehouse = wfmWarehouse,material_type=wf_material)
                # 生成一个窗框对象，将窗框添加窗户到上
                window.set_window_frame(window_frame)
                # 生成一个玻璃对象，将玻璃添加到窗户上
                glass = Glass(area=window_area*(1-wfa_ratio),warehouse = gmWarehouse,material_type=gmWarehouse.get_random_material())
                window.set_glass(glass)
                wall.reset_window(window)
                print("window_features[0] = ",window_features[0], "wall.window._area ",wall._window._area )

    return house

def population_to_exl(x, y):

    x = np.vstack(x)
    y = np.vstack(y)
    result = np.concatenate((x, y), axis=1)
    # 将二维数组转换为 DataFrame
    df = pd.DataFrame(result)
    start_time = datetime.now()  # 设置开始时间
    time_format = start_time.strftime("%Y%m%d-%H%M%S-%f")[:-3]  # 格式化时间，精确到100毫秒
    # 生成带时间序列的文件名
    file_name = f"./predicted_data/test_population/tp_{time_format}.xlsx"
    # 将 DataFrame 写入 Excel 文件
    df.to_excel(file_name, index=False)


# 把数据集导出到 Excel 表中
def export_to_excel(houses_x, houses_y, ds_size,path):
    print("\rExporting the training dataset to Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Dataset"
    #ws.append(ds_bp.get_dataset_headers())
    for i, house_x in enumerate(houses_x):
        if i >= ds_size:
            break

        house_x = house_x.numpy()
        house_y = houses_y[i].numpy()
        row = list(house_x.flatten()) + list(house_y.flatten())
        ws.append(row)
        sys.stdout.write("\rExporting to Excel: %d/%d" % (i+1, len(houses_x)))
        sys.stdout.flush()
    wb.save(path)
    print("\nDone.")

# 从 Excel 表中导入数据集
'''
def import_from_excel(path):
    print("Importing the training dataset from Excel...")
    wb = load_workbook(path)
    ws = wb.active
    houses_x = []
    houses_y = []
    for i, row in enumerate(ws.rows):
        if i == 0:
            continue
        houses_x.append(list(map(lambda x: x.value, row[0:House.HOUSE_FEATURES])))
        houses_y.append(list(map(lambda x: x.value, row[House.HOUSE_FEATURES:])))
        sys.stdout.write("\rImporting from Excel: %d/%d" % (i, ws.max_row))
        sys.stdout.flush()
    print("\nDone.")
    #print("Converting the training dataset to tensors...")
    #houses_x = tf.convert_to_tensor(houses_x, dtype=tf.float64)
    #houses_y = tf.convert_to_tensor(houses_y, dtype=tf.float64)

    #print("houses_x ",houses_x.shape, "houses_y ",houses_y.shape)
    return houses_x, houses_y
    '''

import pandas as pd

def import_from_excel(path):
    print("Importing the training dataset from Excel...")
    df = pd.read_excel(path)
    houses_x = []
    houses_y = []

    for i, row in df.iterrows():
        houses_x.append(row[0:House.HOUSE_FEATURES].tolist())
        houses_y.append(row[House.HOUSE_FEATURES:].tolist())
        sys.stdout.write("\rImporting from Excel: %d/%d" % (i+1, len(df)))
        sys.stdout.flush()

    print("\nDone.")

    #houses_x = tf.convert_to_tensor(houses_x, dtype=tf.float64)
    #houses_y = tf.convert_to_tensor(houses_y, dtype=tf.float64)

    #print("houses_x ", houses_x.shape, "houses_y ", houses_y.shape)

    return houses_x, houses_y


def generate_excel_filename(input_file=None):
    # 获取当前的日期和时间
    #now = datetime.datetime.now()
    now = datetime.now()

    # 将日期和时间转换为字符串
    timestamp = now.strftime('%Y%m%d%H%M%S')

    # 生成一个随机数
    random_num = random.randint(1000, 9999)

    # 创建文件名
    filename = f"{os.path.basename(input_file)}_{timestamp}_{random_num}.xlsx"

    # 创建文件路径
    filepath = os.path.join('./predicted_data/', filename)

    return filepath

def generate_cn_filename(file_path):
    try:
        # 获取文件名和扩展名
        file_name, file_ext = os.path.splitext(file_path)

        # 检查文件名中是否包含'_cn'
        if '_cn' in file_name:
            return file_path

        # 构建新的文件名
        new_file_name = file_name + "_cn" + file_ext

        return new_file_name
    except ValueError as e:
        print("文件名格式错误:", e)
        return None
    except Exception as e:
        print("发生未知错误:", e)
        return None

if __name__ == "__main__":
    x,y = generate_dataset(100)
    export_to_excel(x, y, 100, "./datasets/training_dataset.xlsx")
    #import_from_excel("./datasets/training_dataset.xlsx")
    '''
    x,y = generate_test_population(50,'./houses_json/training_dataset_json/20230611-154118-562269/20230611_154221364406.json')
    population_to_exl(x,y)
    x = np.vstack(x)
    y = np.vstack(y)
    result = np.concatenate((x, y), axis=1)
    np.set_printoptions(threshold=np.inf)
    print("result ---", result)

    # 将二维数组转换为 DataFrame
    df = pd.DataFrame(result)

    # 将 DataFrame 写入 Excel 文件
    df.to_excel('output.xlsx', index=False)
    '''
    # 将Tensor转换为NumPy数组
    #x_np = x.numpy()
    #y_np = y.numpy()

    # 使用NumPy的打印函数打印完整的数组
    #np.set_printoptions(threshold=np.inf)  # 设置打印阈值为无穷大，显示所有元素
    #print(x_np)
    #print(y_np)
